#!/usr/bin/python3.6

import json
from mail import sendmail
from openpyxl import load_workbook

with open("Mail_Config", "r") as file:
    MailDetails = json.load(file)
DetailsXl = MailDetails["Path_XL"]
sub = MailDetails["DefaultSubject"]
folder_name = MailDetails["folder_name"]
message = ""

main_workbook = load_workbook(filename=DetailsXl)
main_sheet = main_workbook.active
for row in main_sheet.iter_rows(min_row=2, values_only=True):
    print(row)
    subject = sub + " |" + row[9]
    file_name = f"sample_{row[9]}.eml"
    message = f"""
  test message for application {row[9]}

  """

    sendmail(Title=subject, receiver_email=row[10], msg=message, folder_save=folder_name, file_name=file_name)
